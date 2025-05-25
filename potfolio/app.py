from flask import Flask, jsonify, request
import sqlite3
import openpyxl

app = Flask(__name__)

# Initialize SQLite database
def init_db():
    conn = sqlite3.connect('portfolio.db')
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS contact_submissions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        email TEXT NOT NULL,
                        message TEXT NOT NULL
                    )''')
    conn.commit()
    conn.close()

# Call the function to initialize the database
init_db()

# Debugging: Log Excel file operations
def log_excel_debug(message):
    with open('excel_debug.log', 'a') as log_file:
        log_file.write(f"{message}\n")

# Function to initialize Excel file
def init_excel():
    try:
        workbook = openpyxl.load_workbook('contact_submissions.xlsx')
        log_excel_debug("Excel file loaded successfully.")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Submissions'
        sheet.append(['Name', 'Email', 'Message'])  # Add headers
        workbook.save('contact_submissions.xlsx')
        log_excel_debug("Excel file created successfully with headers.")

# Call the function to initialize the Excel file
init_excel()

# Home route
@app.route('/')
def home():
    return jsonify({"message": "Welcome to the Flask Backend!"})

# Example API route
@app.route('/api/example', methods=['GET'])
def example_api():
    return jsonify({"data": "This is an example API response."})

# Projects API
@app.route('/api/projects', methods=['GET'])
def get_projects():
    projects = [
        {"id": 1, "title": "E-commerce Security Solution", "description": "A project focused on securing e-commerce platforms."},
        {"id": 2, "title": "Financial Security System", "description": "A system to ensure financial data security."},
        {"id": 3, "title": "Government Security Consulting", "description": "Consulting services for government security."},
        {"id": 4, "title": "Healthcare Vulnerability Audit", "description": "Auditing healthcare systems for vulnerabilities."}
    ]
    return jsonify(projects)

# Contact Form API (Save data to database and Excel)
@app.route('/api/contact', methods=['POST'])
def contact():
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    message = data.get('message')

    # Save to SQLite database
    conn = sqlite3.connect('portfolio.db')
    cursor = conn.cursor()
    cursor.execute('INSERT INTO contact_submissions (name, email, message) VALUES (?, ?, ?)', (name, email, message))
    conn.commit()
    conn.close()

    # Save to Excel file with logging
    try:
        workbook = openpyxl.load_workbook('contact_submissions.xlsx')
        sheet = workbook['Submissions']
        sheet.append([name, email, message])
        workbook.save('contact_submissions.xlsx')
        log_excel_debug(f"Data saved to Excel: {name}, {email}, {message}")
    except Exception as e:
        log_excel_debug(f"Error saving to Excel: {e}")

    return jsonify({"status": "success", "message": "Thank you for reaching out!"})

# API to fetch contact submissions
@app.route('/api/contact-submissions', methods=['GET'])
def get_contact_submissions():
    conn = sqlite3.connect('portfolio.db')
    cursor = conn.cursor()
    cursor.execute('SELECT name, email, message FROM contact_submissions')
    submissions = [
        {"name": row[0], "email": row[1], "message": row[2]} for row in cursor.fetchall()
    ]
    conn.close()
    return jsonify(submissions)

# About Me API
@app.route('/api/about', methods=['GET'])
def about():
    about_info = {
        "name": "Your Name",
        "bio": "A passionate developer with expertise in web security and development.",
        "skills": ["Python", "Flask", "JavaScript", "HTML", "CSS"]
    }
    return jsonify(about_info)

# Run the Flask app
if __name__ == '__main__':
    app.run(debug=True)
